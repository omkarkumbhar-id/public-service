import openpyxl

# WORKBOOK_LOCATION = 'input/tryout-openpyxl.xlsx'
WORKBOOK_LOCATION = 'input/unstructured-data.xlsx'
SHEET_NAME = 'Sheet1'


def read_from_cell(row_index, column_index, read_from_sheet):
    value = read_from_sheet.cell(row=row_index, column=column_index).value
    return value if value else ''


def write_to_cell(row_index, column_index, value, write_to_sheet):
    write_to_sheet.cell(row=row_index, column=column_index).value = value


def get_end_row_index():
    return sheet.max_row + 1


class Table():
    table_sheet: openpyxl.worksheet.worksheet
    table_start_cell: tuple[int, int]
    table_end_cell: tuple[int, int]
    client_name_cell: tuple[int, int]
    client_name: str
    project_name_cell: tuple[int, int]
    project_name: str
    table_columns: list[str] = ['Item', 'Quantity', 'Unit Price', 'Total']
    table_rows: list[list[str]]
    quantity_total: int
    grand_total: int

    def __init__(self, item_row_index, read_from_sheet):
        self.table_sheet = read_from_sheet
        self.table_start_cell = (item_row_index, 1)
        self.client_name_cell = (item_row_index - 4, 1)
        self.project_name_cell = (item_row_index - 1, 1)
        self.client_name = read_from_cell(
            *self.client_name_cell, self.table_sheet)
        self.project_name = read_from_cell(
            *self.project_name_cell, self.table_sheet)

        # Calculating table end cell and total number of rows
        for row_index in range(item_row_index, read_from_sheet.max_row + 1):
            if not (read_from_cell(row_index, 1, read_from_sheet)) and not (read_from_cell(row_index, 3, read_from_sheet)):
                self.table_end_cell = (row_index - 1, 4)
                self.total_number_of_rows = row_index - item_row_index
                break

        # Extracting table rows
        self.table_rows = []
        for row_index in range(self.table_start_cell[0] + 1, self.table_end_cell[0] + 1):
            row = []
            for column_index in range(1, 5):
                row.append(read_from_cell(
                    row_index, column_index, self.table_sheet))
            self.table_rows.append(row)

        self.quantity_total = read_from_cell(
            self.table_end_cell[0] + 1, 2, self.table_sheet)
        self.grand_total = read_from_cell(
            self.table_end_cell[0] + 1, 4, self.table_sheet)

    def print_table(self):
        print()
        # print(f'Client Name: {self.client_name}')
        # print(f'Project Name: {self.project_name}')
        # print(f'Table starts at: {self.table_start_cell}')
        # print(f'Table ends at: {self.table_end_cell}')
        # print(f'Quantity Total: {self.quantity_total}')
        # print(f'Grand Total: {self.grand_total}')
        print('---')
        # Print table
        print(
            f'{"SrNo. ":<10}{"Client Name":<20}{"Project Name":<20}{self.table_columns[0]:<100}{self.table_columns[1]:<15}{self.table_columns[2]:<15}{self.table_columns[3]:<15}')
        print()
        for index, row in enumerate(self.table_rows, start=1):
            print(
                f'{index:<10}{self.client_name:<30}{self.project_name:<30}{row[0]:<100}{row[1]:<15}{row[2]:<15}{row[3]:<15}')
        print('---')


if __name__ == '__main__':
    # start of every script
    workbook = openpyxl.load_workbook(WORKBOOK_LOCATION)
    sheet = workbook[SHEET_NAME]

    # extract data
    tables: list[Table] = []
    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == 'Item':
            tables.append(Table(row_index, sheet))

    # print data
    for table in tables:
        table.print_table()

    # end of every script
    # workbook.save(WORKBOOK_LOCATION)
